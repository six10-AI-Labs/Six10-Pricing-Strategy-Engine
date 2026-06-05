"""
data_loader.py — Load and normalize all data sources for the Pricing Strategy Engine.

All ASIN columns normalized to uppercase strings.
All brand keys normalized via config.brand_name_mapping.
FBA inventory aggregated by brand+asin (multiple SKUs per ASIN collapsed).
Sellerise columns renamed to canonical code-safe names.
Overlapping Sellerise files deduplicated on brand+asin+date.
"""

import os
import glob
import logging
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import numpy as np
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Canonical column name map for Sellerise XLSX files.
# Only top-level aggregate columns are mapped; sub-ad-type columns are ignored.
# ---------------------------------------------------------------------------
SELLERISE_COLUMN_MAP = {
    "ASIN": "asin",
    "Title": "title",
    "Date": "date",
    "Sales": "sales",
    "Product sales": "product_sales",
    "Net profit": "net_profit",
    "Margin": "margin_pct",
    "ROI": "roi",
    "Taxes": "taxes",
    "FBA fees": "fba_fees",
    "Referral fees": "referral_fees",
    "Promo fees": "promo_fees",
    "CoG": "cog",
    "Manual expenses": "manual_expenses",
    "Manual incomes": "manual_incomes",
    "Refunds $": "refunds_dollars",
    "Refunds qty": "refunds_qty",
    "Refund rate %": "refund_rate_pct",
    "Other AMZ transactions": "other_amz_transactions",
    "Ad. cost": "ad_cost",
    "Org. orders": "org_orders",
    "Org. units": "org_units",
    "Org. units %": "org_units_pct",
    "Orders": "orders",
    "Units": "units",
    "MCF orders": "mcf_orders",
    "MCF units": "mcf_units",
    "Coupons": "coupons",
    "Promo": "promo",
    "SNS": "sns",
    "Sessions": "sessions",
    "Page views": "page_views",
    "Sessions mobile": "sessions_mobile",
    "Page views mobile": "page_views_mobile",
    "Conversion": "conversion_pct",
    "True conversion": "true_conversion_pct",
    "Unit session %": "unit_session_pct",
    "TACoS": "tacos_pct",
    "Sales / Unit": "sales_per_unit",
    "Net profit / Unit": "net_profit_per_unit",
    "Refund / Unit": "refund_per_unit",
    "Ad. Cost / Unit": "ad_cost_per_unit",
    "Ad. Sales / Unit": "ad_sales_per_unit",
    "Prod. sales / Order": "prod_sales_per_order",
    "Prod. sales / Unit": "prod_sales_per_unit",
    "Storage fees": "storage_fees",
    "CTR": "ctr_pct",
    "Ad. units %": "ad_units_pct",
}

# Columns to cast to numeric (coerce errors → NaN)
SELLERISE_NUMERIC_COLS = [
    "sales", "product_sales", "net_profit", "margin_pct", "roi", "taxes",
    "fba_fees", "referral_fees", "promo_fees", "cog", "refunds_dollars",
    "refunds_qty", "refund_rate_pct", "ad_cost", "org_orders", "org_units",
    "org_units_pct", "orders", "units", "sessions", "page_views",
    "sessions_mobile", "page_views_mobile", "conversion_pct",
    "true_conversion_pct", "unit_session_pct", "tacos_pct",
    "sales_per_unit", "net_profit_per_unit", "prod_sales_per_unit",
    "storage_fees", "ctr_pct", "ad_units_pct",
]


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

# Return reason classification sets (used by load_return_reasons)
_RETURN_PRICE_REASONS = {"NOT_AS_DESCRIBED", "FOUND_BETTER_PRICE"}


def _clean_product_name(name: str) -> str:
    """Normalize a COGS product name entered in camelCase or hyphen-joined format.

    Handles patterns commonly found in the COGS sheet:
      CalciumIncreaser32oz    → Calcium Increaser 32oz
      PHIncreaser-5lbs        → PH Increaser 5lbs
      PoolOpener-32oz         → Pool Opener 32oz
      RClassicBlue-8oz        → R Classic Blue 8oz
      AlkalinityBooster-2Pack → Alkalinity Booster 2 Pack

    Already-clean names (e.g. "Spa Defoamer - 1 Pint") are returned unchanged.
    """
    if not name or not isinstance(name, str):
        return name
    s = name.strip()
    # 1. Split uppercase run before Uppercase+lowercase: "PHIncreaser" → "PH Increaser"
    #    Also handles single capital before camelCase word: "RClassic" → "R Classic"
    s = re.sub(r'([A-Z])([A-Z][a-z])', r'\1 \2', s)
    # 2. Split lowercase-to-uppercase camelCase: "CalciumIncreaser" → "Calcium Increaser"
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
    # 3. Replace hyphen-as-word-joiner with space; leave " - " separators untouched
    #    (hyphens surrounded by spaces are not replaced)
    s = re.sub(r'(?<!\s)-(?!\s)', ' ', s)
    # 4. Split letter→digit boundary: "Increaser32oz" → "Increaser 32oz"
    s = re.sub(r'([A-Za-z])(\d)', r'\1 \2', s)
    # 5. Split digit→uppercase boundary: "2Pack" → "2 Pack"
    s = re.sub(r'(\d)([A-Z])', r'\1 \2', s)
    # 6. Collapse multiple spaces
    s = re.sub(r' +', ' ', s)
    return s.strip()


def _strip_bom_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Strip UTF-8 BOM artifacts (mangled by cp1252 read) and quotes from column names.

    When a UTF-8-BOM file is read with cp1252, the BOM (0xEF 0xBB 0xBF) appears
    as 'ï»¿' at the start of the first column name.  This helper removes that
    prefix *and* any surrounding quotes from every column so downstream code can
    match columns by simple string equality.
    """
    new_cols = []
    for c in df.columns:
        # Strip BOM as latin-1/cp1252 mangling ('ï»¿') and unicode BOM variants
        c = str(c).lstrip("ï»¿\ufeff").strip('"').strip("'").strip()
        new_cols.append(c)
    df.columns = new_cols
    return df


def _dedup_headers(raw_headers: list) -> list:
    """Given a list of header strings (may have duplicates), return a list
    where duplicates are suffixed _0, _1, _2, ... in order of appearance.
    None/empty headers get placeholder names col_N."""
    counts: dict = defaultdict(int)
    seen: dict = defaultdict(int)
    result = []
    for i, h in enumerate(raw_headers):
        h = str(h).strip() if h is not None else f"col_{i}"
        if h == "None" or h == "":
            h = f"col_{i}"
        if counts[h] > 0:
            result.append(f"{h}_{seen[h]}")
            seen[h] += 1
        else:
            result.append(h)
        counts[h] += 1
        if counts[h] == 2:
            # Rename the FIRST occurrence retroactively
            first_idx = result.index(h) if h in result else -1
            if first_idx >= 0 and result[first_idx] == h:
                result[first_idx] = f"{h}_0"
            seen[h] = 1
    return result


def _load_xlsx_dedup_headers(path: str) -> pd.DataFrame:
    """Load an XLSX file with openpyxl, deduplicate headers, return DataFrame."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    headers = _dedup_headers(list(rows[0]))
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return df


def _normalize_asin(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.upper()


def apply_brand_mapping(value: str, mapping: dict) -> str:
    """Map a raw brand label to a config brand key. Returns original if not found."""
    result = mapping.get(str(value).strip())
    if result is None:
        logger.warning("No brand mapping for '%s' — using as-is", value)
        return str(value).strip().lower().replace(" ", "_")
    return result


# ---------------------------------------------------------------------------
# Sellerise loaders
# ---------------------------------------------------------------------------

def load_sellerise_brand(brand_key: str, config: dict) -> pd.DataFrame:
    """Load all Sellerise XLSX files for one brand folder.

    Returns a DataFrame with canonical column names, deduplicated on
    brand+asin+date (most recently modified file wins on conflict).
    """
    base = config["data"]["base_dir"]
    folder_name = config["data"]["brand_folders"].get(brand_key)
    if not folder_name:
        logger.warning("No folder mapping for brand_key '%s'", brand_key)
        return pd.DataFrame()

    folder_path = Path(base) / "SellerRise Sales Data" / folder_name
    if not folder_path.exists():
        logger.warning("Sellerise folder not found: %s", folder_path)
        return pd.DataFrame()

    xlsx_files = sorted(folder_path.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning("No XLSX files in %s", folder_path)
        return pd.DataFrame()

    frames = []
    for xlsx_path in xlsx_files:
        try:
            df = _load_xlsx_dedup_headers(str(xlsx_path))
        except Exception as e:
            logger.error("Failed to load %s: %s", xlsx_path, e)
            continue
        if df.empty:
            continue
        # Rename to canonical names
        df = df.rename(columns=SELLERISE_COLUMN_MAP)
        # Normalize ASIN
        if "asin" in df.columns:
            df["asin"] = _normalize_asin(df["asin"])
        # Parse date
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
        # Cast numerics
        for col in SELLERISE_NUMERIC_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        # Record source metadata
        df["brand"] = brand_key
        df["file_mtime"] = os.path.getmtime(str(xlsx_path))
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    # Drop rows without ASIN or date
    combined = combined.dropna(subset=["asin", "date"])
    # Dedup on brand+asin+date — keep most recent file
    combined = combined.sort_values(["brand", "asin", "date", "file_mtime"])
    combined = combined.drop_duplicates(subset=["brand", "asin", "date"], keep="last")
    combined = combined.drop(columns=["file_mtime"], errors="ignore")
    combined = combined.reset_index(drop=True)
    return combined


def load_all_sellerise(config: dict) -> pd.DataFrame:
    """Load Sellerise data for all 5 brands and concatenate."""
    brand_keys = list(config["data"]["brand_folders"].keys())
    frames = []
    for brand_key in brand_keys:
        df = load_sellerise_brand(brand_key, config)
        if not df.empty:
            frames.append(df)
            logger.info("Loaded Sellerise %s: %d rows", brand_key, len(df))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


# ---------------------------------------------------------------------------
# COGS loader
# ---------------------------------------------------------------------------

def load_cogs(config: dict) -> pd.DataFrame:
    """Load COGS Sheet.xlsx.

    Returns DataFrame with columns: asin, brand_key, cogs.
    Incomplete coverage is expected (esp. AquaDoc) — gaps filled later.
    """
    path = Path(config["data"]["base_dir"]) / "COGS Sheet.xlsx"
    if not path.exists():
        logger.error("COGS Sheet not found: %s", path)
        return pd.DataFrame(columns=["asin", "brand_key", "cogs"])

    df = _load_xlsx_dedup_headers(str(path))
    if df.empty:
        return pd.DataFrame(columns=["asin", "brand_key", "cogs"])

    # Find the relevant columns (handle slight name variation)
    col_map = {}
    for col in df.columns:
        lc = col.lower().strip()
        if "child" in lc and "asin" in lc:
            col_map["asin"] = col
        elif lc == "brand":
            col_map["brand"] = col
        elif lc == "cogs":
            col_map["cogs"] = col

    if "asin" not in col_map:
        logger.error("Could not find ASIN column in COGS Sheet")
        return pd.DataFrame(columns=["asin", "brand_key", "cogs", "product_name", "pack"])

    # Also capture Product Name and Pack for display in email
    for col in df.columns:
        lc = col.lower().strip()
        if "product" in lc and "name" in lc:
            col_map["product_name"] = col
        elif lc == "pack":
            col_map["pack"] = col

    rename_map = {
        col_map.get("asin"): "asin",
        col_map.get("brand", "Brand"): "brand_raw",
        col_map.get("cogs", "COGS"): "cogs",
    }
    if "product_name" in col_map:
        rename_map[col_map["product_name"]] = "product_name"
    if "pack" in col_map:
        rename_map[col_map["pack"]] = "pack"

    df = df.rename(columns=rename_map)

    df["asin"] = _normalize_asin(df["asin"])
    df["cogs"] = pd.to_numeric(df["cogs"], errors="coerce")
    df = df.dropna(subset=["asin", "cogs"])

    if "product_name" not in df.columns:
        df["product_name"] = None
    if "pack" not in df.columns:
        df["pack"] = None

    # Normalise product names: fix camelCase / hyphen-joined entries from COGS sheet
    if "product_name" in df.columns:
        df["product_name"] = df["product_name"].apply(
            lambda x: _clean_product_name(str(x)) if pd.notna(x) and str(x).strip() not in ("None", "nan", "") else x
        )

    mapping = config.get("brand_name_mapping", {})
    df["brand_key"] = df["brand_raw"].apply(
        lambda v: apply_brand_mapping(str(v), mapping) if pd.notna(v) else "unknown"
    )

    return df[["asin", "brand_key", "cogs", "product_name", "pack"]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Product Families loader
# ---------------------------------------------------------------------------

def load_product_families(config: dict) -> pd.DataFrame:
    """Load Product Families.xlsx — brand derived from sheet name."""
    path = Path(config["data"]["base_dir"]) / "Product Families.xlsx"
    if not path.exists():
        logger.warning("Product Families not found: %s", path)
        return pd.DataFrame()

    mapping = config.get("brand_name_mapping", {})
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    frames = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in ("legend & logic", "legend", "logic"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = _dedup_headers(list(rows[0]))
        data = rows[1:]
        df = pd.DataFrame(data, columns=headers)

        # Find ASIN column
        asin_col = None
        cat_col = None
        family_col = None
        for col in df.columns:
            lc = col.lower().strip()
            if "asin" in lc:
                asin_col = col
            elif "category" in lc:
                cat_col = col
            elif "family" in lc:
                family_col = col

        if asin_col is None:
            logger.warning("No ASIN column in sheet '%s', skipping", sheet_name)
            continue

        df = df.rename(columns={
            asin_col: "asin",
            **({"product_category": cat_col} if cat_col else {}),
            **({"product_family": family_col} if family_col else {}),
        })
        if cat_col:
            df = df.rename(columns={cat_col: "product_category"})
        if family_col:
            df = df.rename(columns={family_col: "product_family"})

        df["asin"] = _normalize_asin(df["asin"])
        df = df.dropna(subset=["asin"])
        df["brand"] = apply_brand_mapping(sheet_name.strip(), mapping)
        frames.append(df)

    wb.close()
    if not frames:
        return pd.DataFrame()
    result = pd.concat(frames, ignore_index=True, sort=False)
    keep_cols = ["asin", "brand"] + [c for c in ["product_category", "product_family"] if c in result.columns]
    return result[keep_cols].reset_index(drop=True)


# ---------------------------------------------------------------------------
# FBA Inventory loader
# ---------------------------------------------------------------------------

def load_fba_inventory(config: dict) -> pd.DataFrame:
    """Load FBA Inventory CSVs for all brands.

    Selects the latest snapshot per brand, then aggregates by brand+asin
    to collapse multiple SKUs into one row.
    """
    base = Path(config["data"]["base_dir"]) / "Amazon FBA Inventory report"
    mapping = config.get("brand_name_mapping", {})

    frames = []
    for brand_folder in base.iterdir():
        if not brand_folder.is_dir():
            continue
        csv_files = list(brand_folder.glob("FBA Inventory*.csv"))
        if not csv_files:
            continue

        # Load all CSVs, then filter to max snapshot-date
        brand_frames = []
        for csv_path in csv_files:
            try:
                df = pd.read_csv(str(csv_path), low_memory=False)
            except Exception as e:
                logger.error("Failed to load FBA CSV %s: %s", csv_path, e)
                continue
            brand_frames.append(df)

        if not brand_frames:
            continue

        combined = pd.concat(brand_frames, ignore_index=True, sort=False)
        # Normalize column names: hyphen → underscore, lowercase
        combined.columns = [c.strip().lower().replace("-", "_").replace(" ", "_")
                            for c in combined.columns]

        if "snapshot_date" not in combined.columns:
            logger.warning("No snapshot_date in FBA CSV for %s", brand_folder.name)
            continue

        combined["snapshot_date"] = pd.to_datetime(combined["snapshot_date"], errors="coerce")
        max_date = combined["snapshot_date"].max()
        combined = combined[combined["snapshot_date"] == max_date].copy()

        if "asin" not in combined.columns:
            logger.warning("No asin column in FBA CSV for %s", brand_folder.name)
            continue

        combined["asin"] = _normalize_asin(combined["asin"])
        # Derive brand key from folder name
        brand_key = apply_brand_mapping(brand_folder.name, mapping)
        combined["brand"] = brand_key

        # Numeric coercion for key columns
        for col in ["your_price", "days_of_supply", "units_shipped_t30",
                    "available", "sales_rank", "units_shipped_t7",
                    "units_shipped_t60", "units_shipped_t90"]:
            if col in combined.columns:
                combined[col] = pd.to_numeric(combined[col], errors="coerce")

        frames.append(combined)

    if not frames:
        logger.error("No FBA inventory data loaded")
        return pd.DataFrame()

    all_fba = pd.concat(frames, ignore_index=True, sort=False)

    # Aggregate by brand+asin to collapse multiple SKUs
    agg_dict = {}
    if "your_price" in all_fba.columns:
        agg_dict["your_price"] = "median"
    if "available" in all_fba.columns:
        agg_dict["available"] = "sum"
    if "units_shipped_t30" in all_fba.columns:
        agg_dict["units_shipped_t30"] = "sum"
    if "units_shipped_t7" in all_fba.columns:
        agg_dict["units_shipped_t7"] = "sum"
    if "days_of_supply" in all_fba.columns:
        # Bug 1 fix: use "max" not "min". min() picks the zero-stock variant
        # (e.g. a retired SKU with 0 available) and flags a healthy ASIN as
        # out of stock. max() reflects the best-stocked fulfillable SKU.
        agg_dict["days_of_supply"] = "max"
    if "sales_rank" in all_fba.columns:
        agg_dict["sales_rank"] = "min"
    if "fba_inventory_level_health_status" in all_fba.columns:
        agg_dict["fba_inventory_level_health_status"] = "first"

    if not agg_dict:
        return all_fba

    aggregated = all_fba.groupby(["brand", "asin"]).agg(agg_dict).reset_index()
    logger.info("FBA inventory loaded: %d unique brand+asin combinations", len(aggregated))
    return aggregated


# ---------------------------------------------------------------------------
# Amazon Business Report loader (AquaDoc only, lagged)
# ---------------------------------------------------------------------------

def load_amazon_business_report(config: dict) -> pd.DataFrame:
    """Load AquaDoc business report (sessions, buy box %).
    This is a lagged, AquaDoc-only supplementary source.
    """
    folder = Path(config["data"]["base_dir"]) / "Amazon Sales and Traffic report"
    csv_files = list(folder.glob("AquaDocBusinessReport*.csv"))
    if not csv_files:
        logger.info("No AquaDoc business report found — skipping")
        return pd.DataFrame()

    # Use the most recently modified file
    csv_files.sort(key=lambda p: os.path.getmtime(str(p)), reverse=True)
    path = csv_files[0]

    try:
        df = pd.read_csv(str(path), low_memory=False)
    except Exception as e:
        logger.error("Failed to load business report %s: %s", path, e)
        return pd.DataFrame()

    # Find child ASIN column
    asin_col = None
    for col in df.columns:
        if "child" in col.lower() and "asin" in col.lower():
            asin_col = col
            break
        elif col.strip().lower() == "(child) asin":
            asin_col = col
            break

    if asin_col is None:
        logger.warning("No (Child) ASIN column in business report")
        return pd.DataFrame()

    df = df.rename(columns={asin_col: "asin"})
    df["asin"] = _normalize_asin(df["asin"])
    df["brand"] = "aquadoc"
    df["report_date"] = datetime.fromtimestamp(os.path.getmtime(str(path))).date()
    df["lagged"] = True

    # Normalize column names
    df.columns = [c.strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("%", "pct")
                  for c in df.columns]

    # Key columns
    numeric_cols = [c for c in df.columns if any(k in c for k in
                    ["session", "buy_box", "unit_session", "page_view", "units_ordered"])]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    logger.info("Business report loaded from %s (lagged, AquaDoc only)", path.name)
    return df


# ---------------------------------------------------------------------------
# NL Brands sub-brand enrichment
# ---------------------------------------------------------------------------

def enrich_nl_brands_sub_brand(sellerise_df: pd.DataFrame, cogs_df: pd.DataFrame) -> pd.DataFrame:
    """Add sub_brand column to NL Brands rows in sellerise_df.

    Priority: COGS Sheet brand → title keywords → nl_brands_other fallback.
    """
    nl_mask = sellerise_df["brand"] == "nl_brands"
    if not nl_mask.any():
        sellerise_df["sub_brand"] = sellerise_df["brand"]
        return sellerise_df

    # Build ASIN→sub_brand map from COGS Sheet
    nl_cogs = cogs_df[cogs_df["brand_key"].isin(["pawmedica", "ruckus", "pure_velvet"])]
    asin_to_sub = dict(zip(nl_cogs["asin"], nl_cogs["brand_key"]))

    def _infer_sub_brand(row):
        if row["brand"] != "nl_brands":
            return row["brand"]
        # Check COGS mapping
        sub = asin_to_sub.get(row["asin"])
        if sub:
            return sub
        # Check title keywords
        title = str(row.get("title", "")).lower()
        if "ruckus" in title:
            return "ruckus"
        if "pawmedica" in title or "paw medica" in title:
            return "pawmedica"
        if "pure velvet" in title or "colostrum" in title:
            return "pure_velvet"
        logger.debug("Could not infer sub_brand for ASIN %s — using nl_brands_other", row["asin"])
        return "nl_brands_other"

    sellerise_df = sellerise_df.copy()
    sellerise_df["sub_brand"] = sellerise_df.apply(_infer_sub_brand, axis=1)
    return sellerise_df


# ---------------------------------------------------------------------------
# Helium10 loader
# ---------------------------------------------------------------------------

def load_helium10(config: dict) -> pd.DataFrame:
    """Load Helium10 keyword-rank and BSR data.

    Sources (merged, latest per ASIN wins):
      1. helium10_snapshot_*.csv  — normalised CSV at root of Helium10 Data/
      2. My Products*.xlsx        — brand-specific files in brand sub-folders

    Returns DataFrame with columns: asin, keyword_avg_rank_trend, category_bsr,
    subcategory_bsr, keyword_avg_rank.
    """
    base = Path(config["data"]["base_dir"]) / "Helium10 Data"
    if not base.exists():
        logger.warning("Helium10 Data folder not found: %s", base)
        return pd.DataFrame()

    frames = []

    # ── 1. Snapshot CSVs ──────────────────────────────────────────────────
    for csv_path in sorted(base.glob("helium10_snapshot_*.csv")):
        try:
            df = pd.read_csv(str(csv_path), low_memory=False)
            df.columns = [c.strip().lower().replace(" ", "_").replace("-", "_")
                          for c in df.columns]
            if "asin" in df.columns:
                df["asin"] = _normalize_asin(df["asin"])
                frames.append(df)
                logger.info("Helium10 snapshot loaded: %s (%d rows)", csv_path.name, len(df))
        except Exception as e:
            logger.error("Failed to load Helium10 snapshot %s: %s", csv_path, e)

    # ── 2. Brand-specific XLSX files (most recent per brand folder) ───────
    _h10_col_map = {
        # XLSX raw names → canonical names
        "asin": "asin",
        "keywords average rank": "keyword_avg_rank",
        "keywords average rank trend": "keyword_avg_rank_trend",
        "keyword average rank": "keyword_avg_rank",
        "keyword average rank trend": "keyword_avg_rank_trend",
        "category bsr": "category_bsr",
        "subcategory bsr": "subcategory_bsr",
    }
    for brand_folder in base.iterdir():
        if not brand_folder.is_dir():
            continue
        xlsx_files = sorted(brand_folder.glob("My Products*.xlsx"))
        if not xlsx_files:
            continue
        xlsx_path = xlsx_files[-1]  # latest by filename sort
        try:
            df = _load_xlsx_dedup_headers(str(xlsx_path))
            if df.empty:
                continue
            # Map known column names (case-insensitive)
            renamed = {}
            for col in df.columns:
                canonical = _h10_col_map.get(col.lower().strip())
                if canonical:
                    renamed[col] = canonical
            df = df.rename(columns=renamed)
            if "asin" not in df.columns:
                continue
            df["asin"] = _normalize_asin(df["asin"])
            frames.append(df)
            logger.debug("Helium10 XLSX loaded: %s (%d rows)", xlsx_path.name, len(df))
        except Exception as e:
            logger.error("Failed to load Helium10 XLSX %s: %s", xlsx_path, e)

    if not frames:
        logger.info("No Helium10 data found")
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)

    # Numeric coercion
    for col in ["keyword_avg_rank_trend", "category_bsr", "subcategory_bsr", "keyword_avg_rank"]:
        if col in combined.columns:
            combined[col] = pd.to_numeric(combined[col], errors="coerce")

    combined = combined.dropna(subset=["asin"])
    combined = combined.drop_duplicates(subset=["asin"], keep="last")

    keep = [c for c in ["asin", "keyword_avg_rank_trend", "category_bsr",
                         "subcategory_bsr", "keyword_avg_rank"] if c in combined.columns]
    combined = combined[keep].reset_index(drop=True)
    logger.info("Helium10 data loaded: %d ASINs", len(combined))
    return combined


# ---------------------------------------------------------------------------
# Fee Preview loader
# ---------------------------------------------------------------------------

def load_fee_preview(config: dict) -> pd.DataFrame:
    """Load Amazon Fee Preview CSVs — per-ASIN exact FBA fulfilment fee.

    Files use cp1252 encoding with a UTF-8 BOM on the first column name.
    _strip_bom_cols() removes the garbled prefix before column matching.

    Returns DataFrame with: asin, brand, fee_preview_fba.
    """
    base = Path(config["data"]["base_dir"]) / "Amazon Fee Preview Report"
    if not base.exists():
        logger.warning("Fee Preview folder not found: %s", base)
        return pd.DataFrame()

    mapping = config.get("brand_name_mapping", {})
    frames = []

    for brand_folder in base.iterdir():
        if not brand_folder.is_dir():
            continue
        csv_files = sorted(brand_folder.glob("*.csv"))
        if not csv_files:
            continue
        csv_path = csv_files[-1]   # most recent by filename sort
        try:
            df = pd.read_csv(str(csv_path), encoding="cp1252", low_memory=False)
            df = _strip_bom_cols(df)
            df.columns = [c.strip().lower().replace("-", "_").replace(" ", "_")
                          for c in df.columns]

            # Locate ASIN column (may be 'asin' or under another key)
            if "asin" not in df.columns:
                asin_col = next((c for c in df.columns if "asin" in c), None)
                if asin_col:
                    df = df.rename(columns={asin_col: "asin"})
                else:
                    logger.warning("No ASIN column in fee preview: %s", csv_path)
                    continue

            df["asin"] = _normalize_asin(df["asin"])
            brand_key = apply_brand_mapping(brand_folder.name, mapping)
            df["brand"] = brand_key

            # Normalise the FBA fee column
            fba_col = next((c for c in df.columns
                            if "fulfillment_fee" in c or "fulfilment_fee" in c), None)
            if fba_col:
                df["fee_preview_fba"] = pd.to_numeric(df[fba_col], errors="coerce")
            else:
                logger.warning("No fulfillment fee column in %s", csv_path)
                continue

            frames.append(df[["asin", "brand", "fee_preview_fba"]].dropna(subset=["asin"]))
            logger.debug("Fee Preview loaded: %s (%d rows)", csv_path.name, len(df))
        except Exception as e:
            logger.error("Failed to load Fee Preview %s: %s", csv_path, e)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combined.drop_duplicates(subset=["brand", "asin"], keep="last")
    combined = combined.reset_index(drop=True)
    logger.info("Fee Preview loaded: %d brand+ASIN entries", len(combined))
    return combined


# ---------------------------------------------------------------------------
# Storage Fees loader
# ---------------------------------------------------------------------------

def load_storage_fees(config: dict) -> pd.DataFrame:
    """Load Amazon Monthly Storage Fees CSVs.

    Computes per-unit monthly storage cost:
        monthly_storage_fee_per_unit = estimated_monthly_storage_fee / average_quantity_on_hand

    Averages across available months when multiple files exist.
    Returns DataFrame with: asin, brand, monthly_storage_fee_per_unit.
    """
    base = Path(config["data"]["base_dir"]) / "Amazon Monthly Storage Fees Report"
    if not base.exists():
        logger.warning("Storage Fees folder not found: %s", base)
        return pd.DataFrame()

    mapping = config.get("brand_name_mapping", {})
    frames = []

    for brand_folder in base.iterdir():
        if not brand_folder.is_dir():
            continue
        csv_files = sorted(brand_folder.glob("*.csv"))
        if not csv_files:
            continue
        brand_key = apply_brand_mapping(brand_folder.name, mapping)

        for csv_path in csv_files:
            try:
                df = pd.read_csv(str(csv_path), encoding="cp1252", low_memory=False)
                df = _strip_bom_cols(df)
                df.columns = [c.strip().lower().replace("-", "_").replace(" ", "_")
                              for c in df.columns]

                # Locate ASIN column
                if "asin" not in df.columns:
                    asin_col = next((c for c in df.columns if "asin" in c), None)
                    if asin_col:
                        df = df.rename(columns={asin_col: "asin"})
                    else:
                        logger.warning("No ASIN column in storage fees: %s", csv_path)
                        continue

                df["asin"] = _normalize_asin(df["asin"])
                df["brand"] = brand_key

                fee_col = next((c for c in df.columns if "monthly_storage_fee" in c
                                or "storage_fee" in c), None)
                qty_col = next((c for c in df.columns if "average_quantity" in c
                                or "avg_quantity" in c), None)

                if fee_col:
                    df["_fee"] = pd.to_numeric(df[fee_col], errors="coerce")
                else:
                    logger.warning("No storage fee column in %s", csv_path)
                    continue

                if qty_col:
                    df["_qty"] = pd.to_numeric(df[qty_col], errors="coerce").replace(0, np.nan)
                    df["monthly_storage_fee_per_unit"] = df["_fee"] / df["_qty"]
                else:
                    # No quantity — treat the fee itself as a per-unit estimate
                    df["monthly_storage_fee_per_unit"] = df["_fee"]

                frames.append(df[["asin", "brand", "monthly_storage_fee_per_unit"]].dropna(subset=["asin"]))
            except Exception as e:
                logger.error("Failed to load Storage Fees %s: %s", csv_path, e)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combined.dropna(subset=["asin", "monthly_storage_fee_per_unit"])

    # Average across months per brand+asin
    agg = (combined.groupby(["brand", "asin"])["monthly_storage_fee_per_unit"]
           .mean().reset_index())
    agg = agg[agg["monthly_storage_fee_per_unit"] > 0]
    logger.info("Storage Fees loaded: %d brand+ASIN entries", len(agg))
    return agg


# ---------------------------------------------------------------------------
# Return Reasons loader
# ---------------------------------------------------------------------------

def load_return_reasons(config: dict) -> pd.DataFrame:
    """Load Amazon FBA Return Reports and classify return reasons.

    Price-related reasons: NOT_AS_DESCRIBED, FOUND_BETTER_PRICE.
    Quality reasons: DEFECTIVE, QUALITY_UNACCEPTABLE, DAMAGED_BY_FC, DAMAGED_BY_CARRIER.
    Choice reasons: UNWANTED_ITEM, ORDERED_WRONG_ITEM.

    Returns DataFrame with: asin, brand, total_returns, price_related_returns,
    price_related_fraction.
    """
    base = Path(config["data"]["base_dir"]) / "Return Reports"
    if not base.exists():
        logger.warning("Return Reports folder not found: %s", base)
        return pd.DataFrame()

    mapping = config.get("brand_name_mapping", {})
    all_rows = []

    for brand_folder in base.iterdir():
        if not brand_folder.is_dir():
            continue
        csv_files = list(brand_folder.glob("*.csv"))
        if not csv_files:
            continue
        brand_key = apply_brand_mapping(brand_folder.name, mapping)

        for csv_path in csv_files:
            try:
                df = pd.read_csv(str(csv_path), encoding="latin-1", low_memory=False)
                if "asin" not in df.columns or "reason" not in df.columns:
                    continue
                df["asin"] = _normalize_asin(df["asin"])
                df["brand"] = brand_key
                if "quantity" in df.columns:
                    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(1)
                else:
                    df["quantity"] = 1.0
                all_rows.append(df[["asin", "brand", "reason", "quantity"]].dropna(subset=["asin"]))
            except Exception as e:
                logger.error("Failed to load Return Report %s: %s", csv_path, e)

    if not all_rows:
        return pd.DataFrame()

    combined = pd.concat(all_rows, ignore_index=True, sort=False)
    combined = combined.dropna(subset=["asin", "reason"])
    combined["reason"] = combined["reason"].str.strip()

    # Aggregate total returns per brand+asin
    total_agg = (combined.groupby(["brand", "asin"])["quantity"]
                 .sum().reset_index().rename(columns={"quantity": "total_returns"}))

    # Aggregate price-related returns
    price_df = combined[combined["reason"].isin(_RETURN_PRICE_REASONS)]
    price_agg = (price_df.groupby(["brand", "asin"])["quantity"]
                 .sum().reset_index().rename(columns={"quantity": "price_related_returns"}))

    result = total_agg.merge(price_agg, on=["brand", "asin"], how="left")
    result["price_related_returns"] = result["price_related_returns"].fillna(0)
    result["price_related_fraction"] = (
        result["price_related_returns"] / result["total_returns"].replace(0, np.nan)
    )
    result = result.dropna(subset=["price_related_fraction"])
    logger.info("Return Reasons loaded: %d brand+ASIN entries", len(result))
    return result


# ---------------------------------------------------------------------------
# 3PL Stock loader
# ---------------------------------------------------------------------------

def load_3pl_stock(config: dict) -> pd.DataFrame:
    """Load 3PL Stock Status CSVs and map SKUs to ASINs via Active SKU list.

    3PL files have SKU not ASIN.  The Active SKU list provides the seller-sku
    → asin mapping.  Latest file per sub-folder is used.

    Returns DataFrame with: asin, tpl_available (units available at 3PL).
    """
    base_3pl = Path(config["data"]["base_dir"]) / "3PL Stock Status report"
    base_sku = Path(config["data"]["base_dir"]) / "Amazon Seller Central Active SKUs list"

    if not base_3pl.exists():
        logger.warning("3PL Stock folder not found: %s", base_3pl)
        return pd.DataFrame()

    # ── Build SKU → ASIN map from all Active SKU files ────────────────────
    sku_to_asin: dict = {}
    if base_sku.exists():
        for sku_file in sorted(base_sku.glob("*.csv")):
            try:
                df = pd.read_csv(str(sku_file), encoding="utf-8-sig", low_memory=False)
                df.columns = [c.strip().lower().replace("-", "_").replace(" ", "_")
                              for c in df.columns]
                # Column may be seller_sku or seller-sku (both handled by replace above)
                sku_col = next((c for c in df.columns if "seller" in c and "sku" in c), None)
                asin_col = "asin" if "asin" in df.columns else None
                if sku_col and asin_col:
                    for _, row in df.iterrows():
                        if pd.notna(row[sku_col]) and pd.notna(row[asin_col]):
                            sku_to_asin[str(row[sku_col]).strip()] = (
                                str(row[asin_col]).strip().upper()
                            )
            except Exception as e:
                logger.error("Failed to load Active SKU file %s: %s", sku_file, e)

    if not sku_to_asin:
        logger.warning("No SKU→ASIN mapping loaded — 3PL stock unavailable")
        return pd.DataFrame()
    logger.debug("SKU→ASIN map: %d entries", len(sku_to_asin))

    # ── Load latest 3PL CSV per sub-folder ────────────────────────────────
    tpl_frames = []
    for folder in base_3pl.iterdir():
        if not folder.is_dir():
            continue
        csv_files = sorted(folder.glob("*.csv"))
        if not csv_files:
            continue
        csv_path = csv_files[-1]   # most recent by filename sort
        try:
            # thousands="," handles comma-formatted numbers like "1,000"
            df = pd.read_csv(str(csv_path), low_memory=False, thousands=",")
            df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
            if "sku" not in df.columns:
                logger.warning("No SKU column in 3PL file %s", csv_path)
                continue
            df["sku"] = df["sku"].astype(str).str.strip()

            avail_col = next((c for c in df.columns if "available" in c), None)
            if avail_col is None:
                logger.warning("No available-stock column in 3PL file %s", csv_path)
                continue

            # Strip commas in case thousands= didn't catch string columns
            df["_avail"] = (
                df[avail_col].astype(str)
                .str.replace(",", "", regex=False)
                .pipe(pd.to_numeric, errors="coerce")
                .fillna(0)
            )
            tpl_frames.append(df[["sku", "_avail"]].dropna(subset=["sku"]))
        except Exception as e:
            logger.error("Failed to load 3PL stock file %s: %s", csv_path, e)

    if not tpl_frames:
        return pd.DataFrame()

    combined = pd.concat(tpl_frames, ignore_index=True, sort=False)
    combined["asin"] = combined["sku"].map(sku_to_asin)
    combined = combined.dropna(subset=["asin"])
    combined["_avail"] = combined["_avail"].fillna(0)

    # Sum available stock across all SKUs for each ASIN
    agg = (combined.groupby("asin")["_avail"].sum()
           .reset_index().rename(columns={"_avail": "tpl_available"}))
    agg = agg[agg["tpl_available"] > 0]
    logger.info("3PL Stock loaded: %d ASINs with buffer stock", len(agg))
    return agg


# ---------------------------------------------------------------------------
# Master dataset builder
# ---------------------------------------------------------------------------

def build_master_dataset(config: dict) -> tuple:
    """Load all data sources and return a tuple:
        (sellerise_df, fba_df, cogs_df, families_df, biz_report_df,
         helium10_df, fee_preview_df, storage_df, return_df, tpl_df)

    sellerise_df includes sub_brand column for NL Brands.
    New sources (helium10, fee preview, storage fees, return reasons, 3PL)
    are optional — empty DataFrames returned if files not present.
    """
    logger.info("Loading Sellerise data...")
    sellerise_df = load_all_sellerise(config)
    logger.info("Total Sellerise rows: %d", len(sellerise_df))

    logger.info("Loading COGS...")
    cogs_df = load_cogs(config)
    logger.info("COGS rows: %d", len(cogs_df))

    logger.info("Loading FBA Inventory...")
    fba_df = load_fba_inventory(config)

    logger.info("Loading Product Families...")
    families_df = load_product_families(config)

    logger.info("Loading Amazon Business Report...")
    biz_report_df = load_amazon_business_report(config)

    logger.info("Loading Helium10 rank data...")
    helium10_df = load_helium10(config)

    logger.info("Loading Fee Preview...")
    fee_preview_df = load_fee_preview(config)

    logger.info("Loading Storage Fees...")
    storage_df = load_storage_fees(config)

    logger.info("Loading Return Reports...")
    return_df = load_return_reasons(config)

    logger.info("Loading 3PL Stock...")
    tpl_df = load_3pl_stock(config)

    # Enrich NL Brands sub-brands
    if not sellerise_df.empty:
        sellerise_df = enrich_nl_brands_sub_brand(sellerise_df, cogs_df)

    return (sellerise_df, fba_df, cogs_df, families_df, biz_report_df,
            helium10_df, fee_preview_df, storage_df, return_df, tpl_df)
