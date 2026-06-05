"""
sheets_writer.py — Write 4-tab Google Sheets dashboard.

Tabs:
  1. Summary      — Top 10 RAISE, Top 10 LOWER, color-coded
  2. Full Catalog — All ASINs, all signals + margin scenarios
  3. History      — Append new recommendations; look-back fills post_* columns
  4. Config       — Created from config.yaml defaults on first run; never overwritten

Color scheme:
  RAISE         → green  #d9ead3
  LOWER         → red    #fce4d6
  LOWER_BLOCKED → orange #ffd966
  HOLD          → white  #ffffff
"""

import logging
import os
from typing import Any, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# ── Optional gspread import ────────────────────────────────────────────────────
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False
    logger.warning("gspread / google-auth not installed — Sheets output disabled")

SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]

# ── Recommendation color map ───────────────────────────────────────────────────
REC_COLORS = {
    "RAISE":                  {"red": 0.851, "green": 0.918, "blue": 0.827},   # #d9ead3 green
    "LOWER":                  {"red": 0.988, "green": 0.894, "blue": 0.839},   # #fce4d6 red-pink
    "LOWER_VELOCITY_DEFENCE": {"red": 1.000, "green": 0.949, "blue": 0.800},   # #fff2cc light amber
    "LOWER_BLOCKED":          {"red": 1.000, "green": 0.851, "blue": 0.400},   # #ffd966 orange
    "HOLD":                   {"red": 1.000, "green": 1.000, "blue": 1.000},   # #ffffff white
    "EXCLUDED":               {"red": 0.941, "green": 0.941, "blue": 0.941},   # #f0f0f0 grey
}

# Custom sort order for Full Catalog — action items first
_REC_SORT_ORDER = {
    "RAISE": 0, "LOWER": 1, "LOWER_VELOCITY_DEFENCE": 2,
    "LOWER_BLOCKED": 3, "HOLD": 4, "EXCLUDED": 5,
}

# ── History tab column schema (D14) ───────────────────────────────────────────
HISTORY_COLUMNS = [
    "run_date", "asin", "title", "brand", "sub_brand",
    "recommendation", "confidence", "composite_score", "reasoning",
    "current_price", "cogs", "cogs_source", "current_margin",
    # Pre-change baselines (captured at run time)
    "pre_conversion", "pre_units_weekly", "pre_net_profit_weekly",
    "pre_margin_pct", "pre_ctr", "pre_refund_rate", "pre_days_of_supply",
    # Team fills after acting (auto-backfilled from Pipeline Actions tab)
    "selected_scenario", "selected_price", "implemented_date", "implemented_by",
    "actual_action_taken", "actual_price_implemented",
    # Margin impact
    "raise_5pct_weekly_impact", "raise_10pct_weekly_impact",
    "raise_5pct_new_price", "raise_10pct_new_price",
    "lower_5pct_scenario_a", "lower_5pct_scenario_b",
    "lower_5pct_break_even_pct", "lower_5pct_viable",
    "lower_5pct_new_price",
    "lower_10pct_scenario_a", "lower_10pct_scenario_b",
    "lower_10pct_break_even_pct", "lower_10pct_viable",
    "lower_10pct_new_price",
    # Look-back results (written by engine on subsequent run)
    "post_conversion", "post_units_weekly", "post_net_profit_weekly", "post_margin_pct",
    "outcome", "revert_flag",
]

SUMMARY_COLUMNS = [
    "asin", "brand", "sub_brand", "title", "product_family", "product_category",
    "recommendation", "confidence", "composite_score", "reasoning",
    "current_price", "current_margin", "avg_weekly_units", "cogs_source", "days_of_supply",
    "raise_5pct_new_price", "raise_5pct_weekly_impact",
    "raise_10pct_new_price", "raise_10pct_weekly_impact",
    "lower_5pct_new_price", "lower_5pct_scenario_a", "lower_5pct_scenario_b",
    "lower_5pct_break_even_pct", "lower_5pct_viable",
    "lower_10pct_new_price", "lower_10pct_scenario_a", "lower_10pct_scenario_b",
    "lower_10pct_break_even_pct", "lower_10pct_viable",
]

FULL_CATALOG_COLUMNS = [
    "asin", "brand", "sub_brand", "effective_brand", "title", "cogs_name",
    "product_family", "product_category",
    "recommendation", "confidence", "composite_score", "reasoning",
    "current_price", "price_source", "cogs", "cogs_source",
    "current_margin", "contribution_margin", "avg_tacos_rate", "margin_floor", "margin_below_floor",
    "avg_referral_rate", "avg_fba_fee_per_unit", "avg_weekly_units",
    "days_of_supply", "sales_rank",
    "ctr_score", "conversion_score", "velocity_score",
    "margin_score", "trend_score", "refund_score", "inventory_score",
    "bsr_score", "yoy_factor",
    "avg_ctr", "brand_avg_ctr", "avg_conversion", "baseline_conversion",
    "avg_units_daily", "avg_refund_rate",
    "raise_5pct_new_price", "raise_5pct_weekly_impact",
    "raise_10pct_new_price", "raise_10pct_weekly_impact",
    "lower_5pct_new_price", "lower_5pct_scenario_a", "lower_5pct_scenario_b",
    "lower_5pct_break_even_pct", "lower_5pct_viable",
    "lower_10pct_new_price", "lower_10pct_scenario_a", "lower_10pct_scenario_b",
    "lower_10pct_break_even_pct", "lower_10pct_viable",
    "exclude",
]


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_client(config: dict) -> Any:
    """Authenticate and return a gspread client."""
    creds_path = os.environ.get("GOOGLE_SHEETS_CREDENTIALS") or config["google_sheets"].get("credentials_path", "")
    if not creds_path or not os.path.exists(creds_path):
        raise FileNotFoundError(
            f"Google Sheets credentials not found at '{creds_path}'. "
            "Set GOOGLE_SHEETS_CREDENTIALS env var."
        )
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    return gspread.authorize(creds)


def _get_or_create_worksheet(spreadsheet, title: str, rows: int = 1000, cols: int = 50):
    """Return existing worksheet or create a new one."""
    try:
        return spreadsheet.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        logger.info("Creating new tab: %s", title)
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def read_excluded_asins_from_sheets(config: dict) -> set:
    """Read (brand, asin) pairs marked exclude=TRUE from the Full Catalog tab.

    Called at the start of each run BEFORE the tab is cleared, so the user's
    checkbox selections survive the rewrite cycle.

    Returns empty set if gspread not available, tab missing, or column absent.
    """
    if not GSPREAD_AVAILABLE:
        return set()

    sheet_id = os.environ.get("GOOGLE_SHEET_ID") or config["google_sheets"].get("sheet_id", "")
    if not sheet_id:
        return set()

    try:
        client = _get_client(config)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            ws = spreadsheet.worksheet("Full Catalog")
        except gspread.exceptions.WorksheetNotFound:
            return set()

        values = ws.get_all_values()
        if not values or len(values) < 2:
            return set()

        header = values[0]
        if "exclude" not in header or "asin" not in header or "brand" not in header:
            return set()

        exc_idx   = header.index("exclude")
        asin_idx  = header.index("asin")
        brand_idx = header.index("brand")

        excluded = set()
        for row in values[1:]:
            if len(row) <= max(exc_idx, asin_idx, brand_idx):
                continue
            exc_val = row[exc_idx].strip().upper()
            if exc_val in ("TRUE", "1", "YES"):
                excluded.add((row[brand_idx].strip(), row[asin_idx].strip()))

        logger.info("Read %d excluded ASINs from Full Catalog", len(excluded))
        return excluded

    except Exception as exc:
        logger.warning("Could not read excluded ASINs from sheet: %s", exc)
        return set()


def read_dismissed_asins_tab(config: dict) -> set:
    """Read (brand, asin) pairs from the 'Dismissed ASINs' tab.

    This tab is populated by the Google Apps Script when team members click
    the '✕ Dismiss' link in the weekly email.  Returns empty set if the tab
    doesn't exist yet (first run before any dismissals).
    """
    if not GSPREAD_AVAILABLE:
        return set()

    sheet_id = os.environ.get("GOOGLE_SHEET_ID") or config["google_sheets"].get("sheet_id", "")
    if not sheet_id:
        return set()

    try:
        client = _get_client(config)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            ws = spreadsheet.worksheet("Dismissed ASINs")
        except gspread.exceptions.WorksheetNotFound:
            return set()   # Tab not yet created — no dismissals logged

        values = ws.get_all_values()
        if not values or len(values) < 2:
            return set()

        header = [h.strip().lower() for h in values[0]]
        if "asin" not in header or "brand" not in header:
            logger.warning("'Dismissed ASINs' tab missing 'brand' or 'asin' column — skipping")
            return set()

        asin_idx  = header.index("asin")
        brand_idx = header.index("brand")

        dismissed = set()
        for row in values[1:]:
            if len(row) <= max(asin_idx, brand_idx):
                continue
            a = row[asin_idx].strip()
            b = row[brand_idx].strip()
            if a:
                dismissed.add((b, a))

        if dismissed:
            logger.info("Read %d dismissed ASIN(s) from 'Dismissed ASINs' tab", len(dismissed))
        return dismissed

    except Exception as exc:
        logger.warning("Could not read 'Dismissed ASINs' tab: %s", exc)
        return set()


def read_pipeline_actions_tab(config: dict) -> pd.DataFrame:
    """Read the 'Pipeline Actions' tab written by the Apps Script form.

    Returns a DataFrame with columns:
      brand, asin, title, recommendation, actioned_date,
      actual_action_taken, actual_price_implemented

    Returns empty DataFrame if the tab doesn't exist yet (no actions logged).
    """
    if not GSPREAD_AVAILABLE:
        return pd.DataFrame()

    sheet_id = os.environ.get("GOOGLE_SHEET_ID") or config["google_sheets"].get("sheet_id", "")
    if not sheet_id:
        return pd.DataFrame()

    try:
        client = _get_client(config)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            ws = spreadsheet.worksheet("Pipeline Actions")
        except gspread.exceptions.WorksheetNotFound:
            return pd.DataFrame()   # Tab not created yet — no actions logged

        values = ws.get_all_values()
        if not values or len(values) < 2:
            return pd.DataFrame()

        # Normalise headers
        raw_headers = [h.strip().lower().replace(" ", "_") for h in values[0]]
        df = pd.DataFrame(values[1:], columns=raw_headers)

        # Standardise column names to what the engine expects
        rename_map = {
            "actioned_date":            "actioned_date",
            "actual_action_taken":      "actual_action_taken",
            "actual_price_implemented": "actual_price_implemented",
        }
        # Apps Script writes "Actual Price Implemented" → normalised above
        df = df.rename(columns=rename_map)

        # Coerce actioned_date to datetime
        if "actioned_date" in df.columns:
            df["actioned_date"] = pd.to_datetime(df["actioned_date"], errors="coerce")

        # Coerce actual_price_implemented to float
        if "actual_price_implemented" in df.columns:
            df["actual_price_implemented"] = pd.to_numeric(
                df["actual_price_implemented"], errors="coerce"
            )

        # Drop rows with no ASIN
        if "asin" in df.columns:
            df = df[df["asin"].str.strip() != ""]

        if not df.empty:
            logger.info("Read %d Pipeline Action(s) from 'Pipeline Actions' tab", len(df))
        return df

    except Exception as exc:
        logger.warning("Could not read 'Pipeline Actions' tab: %s", exc)
        return pd.DataFrame()


def backfill_implemented_dates_from_pipeline(
    history_df: pd.DataFrame,
    pipeline_actions_df: pd.DataFrame,
) -> pd.DataFrame:
    """Back-fill implemented_date, actual_action_taken, actual_price_implemented
    into History rows from the Pipeline Actions tab.

    When a team member clicks '✓ Actioned' in the email and submits the form,
    the action is logged to 'Pipeline Actions' with a date and notes.  This
    function finds the most recent History row per (brand, asin) where
    implemented_date is blank and fills it in — so the look-back in
    run_lookback() can trigger automatically without the team having to
    manually update the History sheet.

    Only the most recent pipeline action per ASIN is used (if multiple exist).
    """
    if pipeline_actions_df.empty or history_df.empty:
        return history_df

    required = {"asin", "actioned_date"}
    if not required.issubset(set(pipeline_actions_df.columns)):
        logger.warning("Pipeline Actions tab missing required columns — skipping backfill")
        return history_df

    history_df = history_df.copy()

    # Ensure implemented_date column exists
    if "implemented_date" not in history_df.columns:
        history_df["implemented_date"] = None
    if "actual_action_taken" not in history_df.columns:
        history_df["actual_action_taken"] = None
    if "actual_price_implemented" not in history_df.columns:
        history_df["actual_price_implemented"] = None

    # Coerce history implemented_date
    history_df["implemented_date"] = pd.to_datetime(
        history_df["implemented_date"], errors="coerce"
    )

    # Build lookup: (brand, asin) → most recent pipeline action row
    pipeline_actions_df = pipeline_actions_df.copy()
    pa_brand_col = "brand" if "brand" in pipeline_actions_df.columns else None

    latest_actions = {}
    for _, pa_row in pipeline_actions_df.iterrows():
        pa_asin  = str(pa_row.get("asin", "")).strip()
        pa_brand = str(pa_row.get("brand", "")).strip() if pa_brand_col else ""
        pa_date  = pa_row.get("actioned_date")
        if not pa_asin or pd.isna(pa_date):
            continue
        key = (pa_brand, pa_asin)
        if key not in latest_actions or pa_date > latest_actions[key]["actioned_date"]:
            latest_actions[key] = pa_row

    # Back-fill History rows
    backfilled = 0
    for idx, hist_row in history_df.iterrows():
        h_asin  = str(hist_row.get("asin", "")).strip()
        h_brand = str(hist_row.get("brand", "")).strip()
        key = (h_brand, h_asin)

        if key not in latest_actions:
            continue

        # Only back-fill if implemented_date is currently blank
        existing_impl = hist_row.get("implemented_date")
        if pd.notna(existing_impl) and str(existing_impl).strip() not in ("", "NaT", "None"):
            continue

        pa = latest_actions[key]
        history_df.at[idx, "implemented_date"]        = pa.get("actioned_date")
        history_df.at[idx, "actual_action_taken"]     = pa.get("actual_action_taken", "")
        history_df.at[idx, "actual_price_implemented"] = pa.get("actual_price_implemented")
        backfilled += 1

    if backfilled:
        logger.info("Backfilled implemented_date for %d History row(s) from Pipeline Actions", backfilled)
    return history_df


def _df_to_values(df: pd.DataFrame) -> list[list]:
    """Convert DataFrame to list-of-lists with None converted to empty string."""
    rows = []
    for _, row in df.iterrows():
        rows.append([
            "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
            if not isinstance(v, (int, float, bool)) else v
            for v in row
        ])
    return rows


def _apply_rec_colors(worksheet, df: pd.DataFrame, header_offset: int = 1):
    """Apply background colors to rows based on recommendation column."""
    if "recommendation" not in df.columns:
        return

    requests = []
    sheet_id = worksheet._properties["sheetId"]

    for i, rec in enumerate(df["recommendation"]):
        color = REC_COLORS.get(str(rec).upper(), REC_COLORS["HOLD"])
        row_idx = i + header_offset + 1  # 0-indexed, +1 for header
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx,
                    "endRowIndex": row_idx + 1,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": color
                    }
                },
                "fields": "userEnteredFormat.backgroundColor",
            }
        })

    if requests:
        worksheet.spreadsheet.batch_update({"requests": requests})


def _freeze_and_bold_header(worksheet):
    """Freeze first row and bold it."""
    sheet_id = worksheet._properties["sheetId"]
    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {"frozenRowCount": 1},
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True}
                    }
                },
                "fields": "userEnteredFormat.textFormat.bold",
            }
        },
    ]
    worksheet.spreadsheet.batch_update({"requests": requests})


def _select_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Return df with only the specified columns (in order), adding missing ones as empty."""
    for col in columns:
        if col not in df.columns:
            df = df.copy()
            df[col] = None
    return df[columns]


def _confidence_rank(conf: str) -> int:
    return {"High": 3, "Medium": 2, "Low": 1}.get(conf, 0)


# ─────────────────────────────────────────────────────────────────────────────
# Tab writers
# ─────────────────────────────────────────────────────────────────────────────

def _write_summary_tab(worksheet, recommendations_df: pd.DataFrame):
    """Write Summary tab: Top 10 RAISE + Top 10 LOWER (excluded ASINs not shown)."""
    df = recommendations_df.copy()
    df["_conf_rank"] = df.get("confidence", pd.Series(dtype=str)).apply(_confidence_rank)

    # Never show excluded ASINs in summary
    active_mask = df.get("exclude", pd.Series(False, index=df.index)).astype(str).str.upper() != "TRUE"
    df = df[active_mask]

    raise_df = (
        df[df["recommendation"] == "RAISE"]
        .sort_values(["_conf_rank", "raise_5pct_weekly_impact"],
                     ascending=[False, False])
        .head(10)
    )
    # Summary shows profitable LOWERs first, then LOWER_VELOCITY_DEFENCE (monitor rows)
    _lower_profitable = (
        df[df["recommendation"] == "LOWER"]
        .sort_values(["_conf_rank", "lower_5pct_scenario_b"], ascending=[False, False])
    )
    _lower_vd = (
        df[df["recommendation"] == "LOWER_VELOCITY_DEFENCE"]
        .sort_values(["_conf_rank", "avg_weekly_units"], ascending=[False, False])
    )
    lower_df = pd.concat([_lower_profitable, _lower_vd]).head(10)

    combined = pd.concat([raise_df, lower_df], ignore_index=True)
    combined = _select_columns(combined, SUMMARY_COLUMNS)

    worksheet.clear()
    header = SUMMARY_COLUMNS
    values = [header] + _df_to_values(combined)
    worksheet.update(values, "A1")
    _freeze_and_bold_header(worksheet)
    _apply_rec_colors(worksheet, combined)
    logger.info(
        "Summary tab written: %d RAISE, %d LOWER, %d LOWER_VD rows",
        len(raise_df), len(_lower_profitable), len(_lower_vd),
    )


def _write_full_catalog_tab(worksheet, recommendations_df: pd.DataFrame):
    """Write Full Catalog tab: all ASINs sorted by action priority (RAISE first, EXCLUDED last)."""
    df = recommendations_df.copy()
    df["_conf_rank"] = df.get("confidence", pd.Series(dtype=str)).apply(_confidence_rank)
    df["_rec_order"] = df["recommendation"].map(_REC_SORT_ORDER).fillna(99)
    df = df.sort_values(["_rec_order", "_conf_rank"], ascending=[True, False])
    df = _select_columns(df, FULL_CATALOG_COLUMNS)

    worksheet.clear()
    header = FULL_CATALOG_COLUMNS
    values = [header] + _df_to_values(df)
    worksheet.update(values, "A1")
    _freeze_and_bold_header(worksheet)
    _apply_rec_colors(worksheet, df)

    # Apply checkbox data validation to the 'exclude' column so it renders as a tickbox
    if "exclude" in FULL_CATALOG_COLUMNS:
        exc_col_idx = FULL_CATALOG_COLUMNS.index("exclude")  # 0-based
        sheet_id = worksheet._properties["sheetId"]
        n_rows = len(df) + 1  # +1 for header
        try:
            worksheet.spreadsheet.batch_update({"requests": [{
                "setDataValidation": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 1,       # skip header
                        "endRowIndex": max(n_rows, 500),
                        "startColumnIndex": exc_col_idx,
                        "endColumnIndex": exc_col_idx + 1,
                    },
                    "rule": {
                        "condition": {"type": "BOOLEAN"},
                        "showCustomUi": True,
                    },
                }
            }]})
        except Exception as exc:
            logger.warning("Could not apply checkbox validation to exclude column: %s", exc)

    logger.info("Full Catalog tab written: %d rows", len(df))


def _write_history_tab(worksheet, recommendations_df: pd.DataFrame, run_date: str):
    """Append new recommendation rows to History tab.

    Existing rows are preserved; new rows are appended.
    The engine does NOT fill post_* or outcome here — that's done by run_lookback()
    before calling write_sheets(), which passes the already-updated history_df.
    """
    # Read existing history rows
    existing_values = worksheet.get_all_values()
    if existing_values:
        existing_df = pd.DataFrame(existing_values[1:], columns=existing_values[0])
    else:
        existing_df = pd.DataFrame(columns=HISTORY_COLUMNS)

    # Build new rows from recommendations_df — skip EXCLUDED ASINs
    new_rows = []
    for _, row in recommendations_df[recommendations_df["recommendation"] != "EXCLUDED"].iterrows():
        new_row = {col: None for col in HISTORY_COLUMNS}
        new_row["run_date"] = run_date
        new_row["asin"] = row.get("asin")
        new_row["title"] = row.get("title")
        new_row["brand"] = row.get("brand")
        new_row["sub_brand"] = row.get("sub_brand")
        new_row["recommendation"] = row.get("recommendation")
        new_row["confidence"] = row.get("confidence")
        new_row["composite_score"] = row.get("composite_score")
        new_row["reasoning"] = row.get("reasoning")
        new_row["current_price"] = row.get("current_price")
        new_row["cogs"] = row.get("cogs")
        new_row["cogs_source"] = row.get("cogs_source")
        new_row["current_margin"] = row.get("current_margin")
        # Pre-change baselines
        new_row["pre_conversion"] = row.get("pre_conversion")
        new_row["pre_units_weekly"] = row.get("pre_units_weekly")
        new_row["pre_net_profit_weekly"] = row.get("pre_net_profit_weekly")
        new_row["pre_margin_pct"] = row.get("pre_margin_pct")
        new_row["pre_ctr"] = row.get("pre_ctr")
        new_row["pre_refund_rate"] = row.get("pre_refund_rate")
        new_row["pre_days_of_supply"] = row.get("pre_days_of_supply")
        # Margin impact columns
        new_row["raise_5pct_weekly_impact"] = row.get("raise_5pct_weekly_impact")
        new_row["raise_10pct_weekly_impact"] = row.get("raise_10pct_weekly_impact")
        new_row["raise_5pct_new_price"] = row.get("raise_5pct_new_price")
        new_row["raise_10pct_new_price"] = row.get("raise_10pct_new_price")
        new_row["lower_5pct_scenario_a"] = row.get("lower_5pct_scenario_a")
        new_row["lower_5pct_scenario_b"] = row.get("lower_5pct_scenario_b")
        new_row["lower_5pct_break_even_pct"] = row.get("lower_5pct_break_even_pct")
        new_row["lower_5pct_viable"] = row.get("lower_5pct_viable")
        new_row["lower_5pct_new_price"] = row.get("lower_5pct_new_price")
        new_row["lower_10pct_scenario_a"] = row.get("lower_10pct_scenario_a")
        new_row["lower_10pct_scenario_b"] = row.get("lower_10pct_scenario_b")
        new_row["lower_10pct_break_even_pct"] = row.get("lower_10pct_break_even_pct")
        new_row["lower_10pct_viable"] = row.get("lower_10pct_viable")
        new_row["lower_10pct_new_price"] = row.get("lower_10pct_new_price")
        # Look-back results default to pending
        new_row["outcome"] = "pending"
        new_row["revert_flag"] = False
        new_rows.append(new_row)

    new_df = pd.DataFrame(new_rows, columns=HISTORY_COLUMNS)

    # Ensure existing_df has all required columns
    for col in HISTORY_COLUMNS:
        if col not in existing_df.columns:
            existing_df[col] = None

    combined = pd.concat([existing_df[HISTORY_COLUMNS], new_df], ignore_index=True)

    worksheet.clear()
    values = [HISTORY_COLUMNS] + _df_to_values(combined)
    worksheet.update(values, "A1")
    _freeze_and_bold_header(worksheet)
    logger.info("History tab written: %d total rows (%d new)", len(combined), len(new_rows))


def _write_history_tab_updated(worksheet, history_df: pd.DataFrame):
    """Write a pre-built history DataFrame (post look-back update) back to the sheet."""
    # Ensure all columns present
    for col in HISTORY_COLUMNS:
        if col not in history_df.columns:
            history_df = history_df.copy()
            history_df[col] = None

    df = history_df[HISTORY_COLUMNS]
    worksheet.clear()
    values = [HISTORY_COLUMNS] + _df_to_values(df)
    worksheet.update(values, "A1")
    _freeze_and_bold_header(worksheet)
    logger.info("History tab updated: %d rows", len(df))


def _write_config_tab(worksheet, config: dict):
    """Populate Config tab from config.yaml defaults. Never overwrites existing data."""
    existing = worksheet.get_all_values()
    if existing and len(existing) > 1:
        logger.info("Config tab already populated — skipping write")
        return

    rows = [["Parameter", "Value", "Notes"]]
    # Signals weights
    for k, v in config["signals"]["weights"].items():
        rows.append([f"signals.weights.{k}", v, "Signal weight (0–1)"])
    # Thresholds
    for k, v in config["signals"]["thresholds"].items():
        rows.append([f"signals.thresholds.{k}", v, ""])
    # Confidence
    for k, v in config["confidence"].items():
        rows.append([f"confidence.{k}", v, ""])
    # Price scenarios
    rows.append(["price_scenarios.raise_pcts", str(config["price_scenarios"]["raise_pcts"]), ""])
    rows.append(["price_scenarios.lower_pcts", str(config["price_scenarios"]["lower_pcts"]), ""])
    rows.append(["price_scenarios.default_elasticity", config["price_scenarios"]["default_elasticity"], ""])
    # Windows
    rows.append(["analysis_window_days", config.get("analysis_window_days", 30), ""])
    rows.append(["prior_window_days", config.get("prior_window_days", 30), ""])
    rows.append(["yoy_window_days", config.get("yoy_window_days", 30), ""])
    rows.append(["lookback_period_days", config.get("lookback_period_days", 14), ""])
    # Brand floors and elasticity
    for brand, bcfg in config["brands"].items():
        rows.append([f"brands.{brand}.margin_floor", bcfg.get("margin_floor", 0.25), ""])
        rows.append([f"brands.{brand}.price_elasticity", bcfg.get("price_elasticity", -0.8), ""])
        rows.append([f"brands.{brand}.seasonal", bcfg.get("seasonal", False), ""])

    worksheet.update(rows, "A1")
    _freeze_and_bold_header(worksheet)
    logger.info("Config tab populated with %d rows", len(rows) - 1)


# ─────────────────────────────────────────────────────────────────────────────
# Public interface
# ─────────────────────────────────────────────────────────────────────────────

def read_history_from_sheets(config: dict) -> pd.DataFrame:
    """Read existing History tab rows. Returns empty DataFrame if tab missing or empty."""
    if not GSPREAD_AVAILABLE:
        return pd.DataFrame(columns=HISTORY_COLUMNS)

    sheet_id = os.environ.get("GOOGLE_SHEET_ID") or config["google_sheets"].get("sheet_id", "")
    if not sheet_id:
        logger.warning("GOOGLE_SHEET_ID not set — skipping history read")
        return pd.DataFrame(columns=HISTORY_COLUMNS)

    try:
        client = _get_client(config)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            ws = spreadsheet.worksheet("History")
        except gspread.exceptions.WorksheetNotFound:
            return pd.DataFrame(columns=HISTORY_COLUMNS)

        values = ws.get_all_values()
        if not values or len(values) < 2:
            return pd.DataFrame(columns=HISTORY_COLUMNS)

        df = pd.DataFrame(values[1:], columns=values[0])
        # Coerce implemented_date to datetime for look-back logic
        if "implemented_date" in df.columns:
            df["implemented_date"] = pd.to_datetime(df["implemented_date"], errors="coerce")
        return df

    except Exception as exc:
        logger.error("Failed to read History tab: %s", exc)
        return pd.DataFrame(columns=HISTORY_COLUMNS)


def write_sheets(
    recommendations_df: pd.DataFrame,
    history_df: Optional[pd.DataFrame],
    config: dict,
    run_date: str,
) -> None:
    """Write all 4 tabs to Google Sheets.

    Args:
        recommendations_df: Full recommendations with margin impact columns merged in.
        history_df: Already look-back-updated history DataFrame (or None to append fresh rows).
        config: Loaded config dict.
        run_date: ISO date string for the run (e.g. "2026-04-08").
    """
    if not GSPREAD_AVAILABLE:
        logger.warning("gspread not available — skipping Sheets write")
        return

    sheet_id = os.environ.get("GOOGLE_SHEET_ID") or config["google_sheets"].get("sheet_id", "")
    if not sheet_id:
        logger.warning("GOOGLE_SHEET_ID not set — skipping Sheets write")
        return

    try:
        client = _get_client(config)
        spreadsheet = client.open_by_key(sheet_id)
    except Exception as exc:
        logger.error("Failed to open Google Sheet: %s", exc)
        return

    # Ensure all 4 tabs exist
    summary_ws = _get_or_create_worksheet(spreadsheet, "Summary", rows=200, cols=len(SUMMARY_COLUMNS) + 5)
    catalog_ws = _get_or_create_worksheet(spreadsheet, "Full Catalog", rows=1000, cols=len(FULL_CATALOG_COLUMNS) + 5)
    history_ws = _get_or_create_worksheet(spreadsheet, "History", rows=5000, cols=len(HISTORY_COLUMNS) + 5)
    config_ws = _get_or_create_worksheet(spreadsheet, "Config", rows=200, cols=10)

    # Write tabs
    _write_summary_tab(summary_ws, recommendations_df)
    _write_full_catalog_tab(catalog_ws, recommendations_df)

    if history_df is not None and not history_df.empty:
        _write_history_tab_updated(history_ws, history_df)
    else:
        _write_history_tab(history_ws, recommendations_df, run_date)

    _write_config_tab(config_ws, config)
    logger.info("Google Sheets write complete")


def write_local_excel(
    recommendations_df: pd.DataFrame,
    history_df: Optional[pd.DataFrame],
    config: dict,
    run_date: str,
    output_path: str,
) -> None:
    """Write all 4 tabs to a local Excel file (for --dry-run / --output-local)."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary
        raise_df = recommendations_df[recommendations_df["recommendation"] == "RAISE"].copy()
        raise_df["_cr"] = raise_df.get("confidence", pd.Series(dtype=str)).apply(_confidence_rank)
        raise_df = raise_df.sort_values(["_cr", "raise_5pct_weekly_impact"], ascending=[False, False]).head(10)

        _lp = recommendations_df[recommendations_df["recommendation"] == "LOWER"].copy()
        _lp["_cr"] = _lp.get("confidence", pd.Series(dtype=str)).apply(_confidence_rank)
        _lp = _lp.sort_values(["_cr", "lower_5pct_scenario_b"], ascending=[False, False])
        _lvd = recommendations_df[recommendations_df["recommendation"] == "LOWER_VELOCITY_DEFENCE"].copy()
        _lvd["_cr"] = _lvd.get("confidence", pd.Series(dtype=str)).apply(_confidence_rank)
        _lvd = _lvd.sort_values(["_cr", "avg_weekly_units"], ascending=[False, False])
        lower_df = pd.concat([_lp, _lvd]).head(10)

        summary_df = pd.concat([raise_df, lower_df], ignore_index=True)
        summary_df = _select_columns(summary_df, SUMMARY_COLUMNS)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Full Catalog
        catalog_df = _select_columns(recommendations_df.copy(), FULL_CATALOG_COLUMNS)
        catalog_df.to_excel(writer, sheet_name="Full Catalog", index=False)

        # History
        if history_df is not None and not history_df.empty:
            hist_out = history_df.copy()
            for col in HISTORY_COLUMNS:
                if col not in hist_out.columns:
                    hist_out[col] = None
            hist_out[HISTORY_COLUMNS].to_excel(writer, sheet_name="History", index=False)
        else:
            # Build new history rows from recommendations_df — skip EXCLUDED
            new_rows = []
            for _, row in recommendations_df[recommendations_df["recommendation"] != "EXCLUDED"].iterrows():
                new_row = {col: None for col in HISTORY_COLUMNS}
                new_row["run_date"] = run_date
                for col in ["asin", "title", "brand", "sub_brand", "recommendation", "confidence",
                            "composite_score", "reasoning", "current_price", "cogs",
                            "cogs_source", "current_margin",
                            "pre_conversion", "pre_units_weekly", "pre_net_profit_weekly",
                            "pre_margin_pct", "pre_ctr", "pre_refund_rate", "pre_days_of_supply",
                            "raise_5pct_weekly_impact", "raise_10pct_weekly_impact",
                            "raise_5pct_new_price", "raise_10pct_new_price",
                            "lower_5pct_scenario_a", "lower_5pct_scenario_b",
                            "lower_5pct_break_even_pct", "lower_5pct_viable", "lower_5pct_new_price",
                            "lower_10pct_scenario_a", "lower_10pct_scenario_b",
                            "lower_10pct_break_even_pct", "lower_10pct_viable", "lower_10pct_new_price"]:
                    new_row[col] = row.get(col)
                new_row["outcome"] = "pending"
                new_row["revert_flag"] = False
                new_rows.append(new_row)
            pd.DataFrame(new_rows, columns=HISTORY_COLUMNS).to_excel(
                writer, sheet_name="History", index=False
            )

        # Config tab
        config_rows = [["Parameter", "Value", "Notes"]]
        for k, v in config["signals"]["weights"].items():
            config_rows.append([f"signals.weights.{k}", v, "Signal weight"])
        for k, v in config["signals"]["thresholds"].items():
            config_rows.append([f"signals.thresholds.{k}", v, ""])
        for k, v in config["confidence"].items():
            config_rows.append([f"confidence.{k}", v, ""])
        config_rows.append(["price_scenarios.raise_pcts", str(config["price_scenarios"]["raise_pcts"]), ""])
        config_rows.append(["price_scenarios.lower_pcts", str(config["price_scenarios"]["lower_pcts"]), ""])
        config_rows.append(["price_scenarios.default_elasticity", config["price_scenarios"]["default_elasticity"], ""])
        config_rows.append(["analysis_window_days", config.get("analysis_window_days", 30), ""])
        config_rows.append(["lookback_period_days", config.get("lookback_period_days", 14), ""])
        for brand, bcfg in config["brands"].items():
            config_rows.append([f"brands.{brand}.margin_floor", bcfg.get("margin_floor", 0.25), ""])
            config_rows.append([f"brands.{brand}.price_elasticity", bcfg.get("price_elasticity", -0.8), ""])
            config_rows.append([f"brands.{brand}.seasonal", bcfg.get("seasonal", False), ""])

        pd.DataFrame(config_rows[1:], columns=config_rows[0]).to_excel(
            writer, sheet_name="Config", index=False
        )

        # ── Apply column widths + disable wrap on long-text columns ────────────
        # openpyxl default is 8.43 chars wide; set sensible widths so the title
        # column shows horizontally rather than wrapping into many row-height lines.
        _apply_local_excel_formatting(writer)

    logger.info("Local Excel written to %s", output_path)


def _apply_local_excel_formatting(writer: "pd.ExcelWriter") -> None:
    """Set column widths and disable wrap-text for all local Excel sheets.

    Column widths are keyed by column header name (case-insensitive).
    Long-text columns (title, reasoning, cogs_name) get wrap_text=False so
    Excel displays them wide rather than collapsing into tall narrow rows.
    """
    try:
        from openpyxl.styles import Alignment
    except ImportError:
        return  # openpyxl unavailable — no-op

    # Explicit widths for known columns (characters)
    _WIDTHS: dict = {
        "title":              58,
        "cogs_name":          45,
        "reasoning":          50,
        "reasoning_short":    50,
        "asin":               14,
        "brand":              16,
        "sub_brand":          16,
        "effective_brand":    16,
        "product_family":     18,
        "product_category":   18,
        "size":               10,
        "recommendation":     18,
        "confidence":         10,
        "composite_score":    10,
        "price_source":       14,
        "cogs_source":        12,
        "run_date":           12,
        "implemented_date":   16,
        "actual_action_taken": 22,
    }
    _DEFAULT_WIDTH   = 13
    # Columns whose cells should NOT wrap (show as a single wide line)
    _NO_WRAP = {"title", "cogs_name", "reasoning", "reasoning_short", "actual_action_taken"}
    _TOP_ALIGN = Alignment(wrap_text=False, vertical="top")

    for ws in writer.sheets.values():
        # Build {col_name_lower: column_letter} from header row (row 1)
        header_map: dict = {}
        for cell in ws[1]:
            if cell.value is not None:
                header_map[str(cell.value).lower()] = cell.column_letter

        # Set column widths
        for col_name_lower, col_letter in header_map.items():
            width = _WIDTHS.get(col_name_lower, _DEFAULT_WIDTH)
            ws.column_dimensions[col_letter].width = width

        # Disable wrap_text on long-text columns so rows stay 1-line tall
        for col_name_lower in _NO_WRAP:
            col_letter = header_map.get(col_name_lower)
            if col_letter is None:
                continue
            col_idx = ws[f"{col_letter}1"].column
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row,
                min_col=col_idx, max_col=col_idx,
            ):
                for cell in row:
                    cell.alignment = _TOP_ALIGN
